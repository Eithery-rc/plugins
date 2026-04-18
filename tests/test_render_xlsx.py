import json
from pathlib import Path

import openpyxl
import pytest

from config import Profile, Contractor
from render_xlsx import render_journal_xlsx


FIXTURE = Path(__file__).parent / "fixtures" / "sample_extracted.json"


def _profile():
    return Profile(
        fio="ИП ТЕСТОВЫЙ",
        inn="000000000000",
        ogrnip="300000000000000",
        tax_system="PSN",
        ruble_account="40802810123456789012",
        bank_bic="044525000",
    )


def _contractor():
    return Contractor(
        name="Blu Banyan Inc.",
        inn="",
        operation_type="Начисление вознаграждения по агентскому договору",
        description_template="поступление средств за {month} {year}",
    )


def _rates():
    return {
        "2025-07-11": 77.9029,
        "2025-07-28": 79.5527,
        "2025-08-15": 79.8100,
        "2025-08-31": 80.5000,
        "2025-09-26": 83.6069,
    }


def _render(tmp_path):
    data = json.loads(FIXTURE.read_text(encoding="utf-8"))
    out = tmp_path / "journal.xlsx"
    render_journal_xlsx(
        extracted=data,
        profile=_profile(),
        contractors={"Blu Banyan Inc.": _contractor()},
        rates=_rates(),
        output_path=out,
    )
    return out


def test_produces_three_sheets(tmp_path):
    out = _render(tmp_path)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Журнал", "ВЭД", "Сводка"]


def test_journal_has_header_and_10_columns(tmp_path):
    out = _render(tmp_path)
    ws = openpyxl.load_workbook(out)["Журнал"]
    header = [c.value for c in ws[1]]
    expected = [
        "Дата", "Контрагент", "Тип документа", "№ документа",
        "Сумма ₽ (Учитывается в патенте)", "Тип операции", "Описание",
        "USD", "Курс ЦБ", "Банк-источник",
    ]
    assert header == expected


def test_journal_only_incoming_rows(tmp_path):
    """Sample has 3 in + 1 out + 1 fee; Журнал = 3 rows (+ header)."""
    out = _render(tmp_path)
    ws = openpyxl.load_workbook(out)["Журнал"]
    assert ws.max_row == 4  # 1 header + 3 data rows


def test_journal_row_values(tmp_path):
    out = _render(tmp_path)
    ws = openpyxl.load_workbook(out)["Журнал"]
    first_data = [c.value for c in ws[2]]
    # sample_extracted is sorted asc by date: first incoming is 2025-07-11
    assert first_data[0] == "11.07.2025"
    assert first_data[1] == "Blu Banyan Inc."
    assert first_data[2] == "Банковский ордер"
    assert first_data[3] == "FDF2507102885800"
    assert first_data[4] == round(4000 * 77.9029, 2)
    assert first_data[5] == "Начисление вознаграждения по агентскому договору"
    assert first_data[6] == "поступление средств за июнь 2025"
    assert first_data[7] == 4000.00
    assert first_data[8] == 77.9029
    assert first_data[9] == "Jusan"


def test_ved_sheet_aggregates_by_vo_code(tmp_path):
    out = _render(tmp_path)
    ws = openpyxl.load_workbook(out)["ВЭД"]
    text = "\n".join(
        "\t".join(str(c.value) if c.value is not None else "" for c in row)
        for row in ws.iter_rows()
    )
    assert "USD" in text
    assert "840" in text
    assert "19157.12" in text  # opening
    assert "26751.79" in text  # closing
    assert "20200" in text  # incoming code
    assert "61100" in text  # own-transfer code
    assert "80150" in text  # fee code
    # 20200 total: 3 × 4000 = 12000
    assert "12000" in text
    # 61100: 2000
    assert "2000" in text
    # 80150: 22.5
    assert "22.5" in text


def test_summary_sheet_has_totals(tmp_path):
    out = _render(tmp_path)
    ws = openpyxl.load_workbook(out)["Сводка"]
    text = "\n".join(
        "\t".join(str(c.value) if c.value is not None else "" for c in row)
        for row in ws.iter_rows()
    )
    # Total incoming RUB = 4000*(77.9029 + 79.5527 + 83.6069)
    expected_total_rub = round(4000 * (77.9029 + 79.5527 + 83.6069), 2)
    assert str(expected_total_rub) in text or f"{expected_total_rub:.2f}" in text
