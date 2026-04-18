"""Render journal.xlsx with three sheets: Журнал, ВЭД, Сводка."""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from config import Profile, Contractor
from render_1c import _to_ddmmyyyy, _payment_description


JOURNAL_HEADERS = [
    "Дата", "Контрагент", "Тип документа", "№ документа",
    "Сумма ₽ (Учитывается в патенте)", "Тип операции", "Описание",
    "USD", "Курс ЦБ", "Банк-источник",
]

VO_DESCRIPTIONS = {
    "20200": "Зачисление от нерезидента за услуги",
    "61100": "Перевод резидентом на свой счёт за пределами РФ",
    "80150": "Комиссии банка-нерезидента",
}


def _bold(cell):
    cell.font = Font(bold=True)


def _write_journal(ws, extracted, profile, contractors, rates):
    """Fill 'Журнал' sheet. Returns total RUB for use in Сводка."""
    for col, h in enumerate(JOURNAL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _bold(cell)

    default_contractor = Contractor(
        name="", inn="", operation_type="",
        description_template="поступление средств за {month} {year}",
    )

    row = 2
    total_rub = 0.0
    for tx in extracted["transactions"]:
        if tx["direction"] != "in":
            continue
        rate = rates[tx["date"]]
        amount_rub = round(tx["amount"] * rate, 2)
        total_rub += amount_rub
        cname = tx["counterparty"]["name"]
        contractor = contractors.get(cname, default_contractor)
        description = _payment_description(tx["date"], contractor.description_template)
        operation_type = contractor.operation_type or "Оплата товаров и услуг клиентами"

        ws.cell(row=row, column=1, value=_to_ddmmyyyy(tx["date"]))
        ws.cell(row=row, column=2, value=cname)
        ws.cell(row=row, column=3, value="Банковский ордер")
        ws.cell(row=row, column=4, value=tx["reference"])
        ws.cell(row=row, column=5, value=amount_rub)
        ws.cell(row=row, column=6, value=operation_type)
        ws.cell(row=row, column=7, value=description)
        ws.cell(row=row, column=8, value=tx["amount"])
        ws.cell(row=row, column=9, value=rate)
        ws.cell(row=row, column=10, value=extracted["bank"])
        row += 1

    # Auto-width columns
    for col_idx, _ in enumerate(JOURNAL_HEADERS, 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, row)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    return total_rub


def _write_ved(ws, extracted):
    """Fill 'ВЭД' sheet — 181-И report."""
    account = extracted["account"]
    txs = extracted["transactions"]

    # Aggregate by direction × vo_code
    by_code: Dict[str, Dict[str, float]] = defaultdict(lambda: {"amount": 0.0, "count": 0})
    for t in txs:
        code = t["vo_code"]
        by_code[code]["amount"] += t["amount"]
        by_code[code]["count"] += 1

    total_in = sum(t["amount"] for t in txs if t["direction"] == "in")
    total_out = sum(t["amount"] for t in txs if t["direction"] in ("out", "fee"))

    _bold(ws.cell(row=1, column=1, value="Отчёт по ВЭД (Инструкция ЦБ РФ 181-И)"))
    ws.cell(row=2, column=1, value="Счёт:"); ws.cell(row=2, column=2, value=account["number"])
    ws.cell(row=3, column=1, value="Валюта:"); ws.cell(row=3, column=2, value=f'{account["iso_code"]} — {account["currency"]}')
    ws.cell(row=4, column=1, value="Период:")
    ws.cell(row=4, column=2, value=f'{_to_ddmmyyyy(extracted["period"]["start"])} — {_to_ddmmyyyy(extracted["period"]["end"])}')
    ws.cell(row=5, column=1, value="Остаток на начало:"); ws.cell(row=5, column=2, value=account["opening_balance"])
    ws.cell(row=6, column=1, value="Остаток на конец:"); ws.cell(row=6, column=2, value=account["closing_balance"])
    ws.cell(row=7, column=1, value="Зачислено всего:"); ws.cell(row=7, column=2, value=total_in)
    ws.cell(row=8, column=1, value="Списано всего:"); ws.cell(row=8, column=2, value=total_out)

    # Code breakdown
    row = 10
    _bold(ws.cell(row=row, column=1, value="Код ВО"))
    _bold(ws.cell(row=row, column=2, value="Описание"))
    _bold(ws.cell(row=row, column=3, value="Сумма"))
    _bold(ws.cell(row=row, column=4, value="Кол-во операций"))
    row += 1
    for code in sorted(by_code.keys()):
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=VO_DESCRIPTIONS.get(code, "(см. 181-И)"))
        ws.cell(row=row, column=3, value=round(by_code[code]["amount"], 2))
        ws.cell(row=row, column=4, value=by_code[code]["count"])
        row += 1

    # Detail per transaction
    row += 1
    _bold(ws.cell(row=row, column=1, value="Детализация операций"))
    row += 1
    _bold(ws.cell(row=row, column=1, value="Дата"))
    _bold(ws.cell(row=row, column=2, value="Направление"))
    _bold(ws.cell(row=row, column=3, value="Сумма"))
    _bold(ws.cell(row=row, column=4, value="Код ВО"))
    _bold(ws.cell(row=row, column=5, value="Контрагент"))
    _bold(ws.cell(row=row, column=6, value="Назначение"))
    _bold(ws.cell(row=row, column=7, value="Reference"))
    row += 1
    for t in txs:
        ws.cell(row=row, column=1, value=_to_ddmmyyyy(t["date"]))
        ws.cell(row=row, column=2, value=t["direction"])
        ws.cell(row=row, column=3, value=t["amount"])
        ws.cell(row=row, column=4, value=t["vo_code"])
        ws.cell(row=row, column=5, value=t["counterparty"].get("name", ""))
        ws.cell(row=row, column=6, value=t.get("purpose", ""))
        ws.cell(row=row, column=7, value=t.get("reference", ""))
        row += 1

    for col_idx in range(1, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def _write_summary(ws, extracted, total_rub, rates):
    """Fill 'Сводка' sheet."""
    _bold(ws.cell(row=1, column=1, value="Сводка периода"))
    ws.cell(row=2, column=1, value="Банк:"); ws.cell(row=2, column=2, value=extracted["bank"])
    ws.cell(row=3, column=1, value="Период:")
    ws.cell(row=3, column=2, value=f'{_to_ddmmyyyy(extracted["period"]["start"])} — {_to_ddmmyyyy(extracted["period"]["end"])}')

    incoming = [t for t in extracted["transactions"] if t["direction"] == "in"]
    total_usd = sum(t["amount"] for t in incoming)
    ws.cell(row=4, column=1, value="Поступлений:"); ws.cell(row=4, column=2, value=len(incoming))
    ws.cell(row=5, column=1, value="Всего USD:"); ws.cell(row=5, column=2, value=total_usd)
    total_rub_rounded = round(total_rub, 2)
    ws.cell(row=6, column=1, value="Всего ₽ (по курсу ЦБ):")
    ws.cell(row=6, column=2, value=str(total_rub_rounded))

    for col_idx in range(1, 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 30


def render_journal_xlsx(
    extracted: dict,
    profile: Profile,
    contractors: Dict[str, Contractor],
    rates: Dict[str, float],
    output_path: Path,
) -> None:
    """Create journal.xlsx with three sheets: Журнал, ВЭД, Сводка."""
    wb = Workbook()
    # Default sheet
    ws_journal = wb.active
    ws_journal.title = "Журнал"
    total_rub = _write_journal(ws_journal, extracted, profile, contractors, rates)

    ws_ved = wb.create_sheet("ВЭД")
    _write_ved(ws_ved, extracted)

    ws_summary = wb.create_sheet("Сводка")
    _write_summary(ws_summary, extracted, total_rub, rates)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
